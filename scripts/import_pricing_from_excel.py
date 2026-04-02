"""
import_pricing_from_excel.py
───────────────────────────
Reads the CIR pricing Excel file and syncs all pricing tables in PostgreSQL.

Run from the project root:
    uv run python scripts/import_pricing_from_excel.py

What it does:
  1. Opens 250630_Database of Indicative System Prices_V2_SS_GB.xlsx
  2. Parses: System Price, Engineering, Permitting, Bonding, Sales & Use Tax sheets
  3. TRUNCATE + INSERT each pricing_* table (full refresh — idempotent)
  4. Leaves benchmark_log, session tables, and other tables untouched
  5. Prints a summary row count per table

Requirements: openpyxl (add to pyproject.toml if not present)
    uv add openpyxl
"""

import sys
import os
import re
from pathlib import Path

# ── Resolve project root ──────────────────────────────────────────────────────
ROOT = Path(__file__).parent.parent
sys.path.insert(0, str(ROOT))

EXCEL_PATH = ROOT / "cir_old_manual_flow_docs" / "capex" / \
             "250630_Database of Indicative System Prices_V2_SS_GB.xlsx"

# ── DB connection ─────────────────────────────────────────────────────────────
def _get_conn():
    from dotenv import load_dotenv
    load_dotenv(ROOT / ".env")
    import psycopg2
    from mycir_agent.config import SESSION_DB_URL
    m = re.match(
        r"postgresql\+psycopg2://([^:]+):([^@]*)@([^:/]+):(\d+)/(.+)",
        SESSION_DB_URL,
    )
    if not m:
        raise ValueError(f"Cannot parse SESSION_DB_URL: {SESSION_DB_URL}")
    user, password, host, port, dbname = m.groups()
    return psycopg2.connect(
        host=host, port=int(port), dbname=dbname,
        user=user, password=password,
    )


# ─────────────────────────────────────────────────────────────────────────────
# SHEET PARSERS
# Each function returns a list of dicts to INSERT.
# Adjust row/column indices if the Excel layout ever changes.
# ─────────────────────────────────────────────────────────────────────────────

def _safe_float(val, default=None):
    try:
        return float(val) if val not in (None, "", "N/A") else default
    except (TypeError, ValueError):
        return default


def parse_system_price(ws) -> list[dict]:
    """
    System Price sheet — rows organised as:
      Col A: System Type (GM / RT / CP)
      Col B: Size min (MWp)
      Col C: Size max (MWp)
      Col D: Module $/Wp
      Col E: Inverter $/Wp
      Col F: Racking $/Wp
      Col G: Racking SAT $/Wp (GM only)
      Col H: BOS $/Wp
      Col I: Mechanical $/Wp
      Col J: Electrical $/Wp
      Col K: Civil $/Wp
      Col L: Overhead $/Wp
      Col M: SGA $/Wp
      Col N: Contingency %  (stored as 0.05 or 5 — normalised below)
      Col O: Margin %

    IMPORTANT: Adjust indices to match actual Excel column layout.
    Run with --dry-run to preview without writing to DB.
    """
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        stype = str(row[0]).strip().upper() if row[0] else None
        if stype not in ("GM", "RT", "CP"):
            continue
        size_min = _safe_float(row[1])
        size_max = _safe_float(row[2])
        if size_min is None or size_max is None:
            continue

        def pct(v):
            # Normalise: if stored as 5 treat as 5%, convert to 0.05
            f = _safe_float(v, 0)
            return f / 100 if f > 1 else f

        rows.append({
            "system_type":      stype,
            "size_min_mwp":     size_min,
            "size_max_mwp":     size_max,
            "module_per_wp":    _safe_float(row[3], 0),
            "inverter_per_wp":  _safe_float(row[4], 0),
            "racking_per_wp":   _safe_float(row[5], 0),
            "racking_sat_per_wp": _safe_float(row[6]),
            "bos_per_wp":       _safe_float(row[7], 0),
            "mechanical_per_wp": _safe_float(row[8], 0),
            "electrical_per_wp": _safe_float(row[9], 0),
            "civil_per_wp":     _safe_float(row[10], 0),
            "overhead_per_wp":  _safe_float(row[11], 0),
            "sga_per_wp":       _safe_float(row[12], 0),
            "contingency_pct":  pct(row[13]),
            "margin_pct":       pct(row[14]),
        })
    return rows


def parse_engineering(ws) -> list[dict]:
    """
    Engineering sheet — fixed cost by MW band.
    Columns: Size min, Size max, Electrical USD, Civil USD, Substation USD
    """
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        lo = _safe_float(row[0])
        hi = _safe_float(row[1])
        if lo is None or hi is None:
            continue
        rows.append({
            "size_min_mwp": lo,
            "size_max_mwp": hi,
            "electrical_usd": _safe_float(row[2], 0),
            "civil_usd":      _safe_float(row[3], 0),
            "substation_usd": _safe_float(row[4], 0),
        })
    return rows


def parse_permitting(ws) -> list[dict]:
    """
    Permitting sheet — fixed cost by MW band with optional itemised breakdown.
    Columns: Size min, Size max, Local Counsel, Environmental, Civil Survey,
             Facilitation, Total
    """
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        lo = _safe_float(row[0])
        hi = _safe_float(row[1])
        if lo is None or hi is None:
            continue
        counsel = _safe_float(row[2], 0)
        enviro  = _safe_float(row[3], 0)
        civil   = _safe_float(row[4], 0)
        facil   = _safe_float(row[5], 0)
        total   = _safe_float(row[6]) or (counsel + enviro + civil + facil)
        rows.append({
            "size_min_mwp":    lo,
            "size_max_mwp":    hi,
            "total_usd":       total,
            "local_counsel_usd": counsel,
            "environmental_usd": enviro,
            "civil_survey_usd":  civil,
            "facilitation_usd":  facil,
        })
    return rows


def parse_bonding(ws) -> list[dict]:
    """
    Bonding sheet — rate % by MW band.
    Columns: Size min, Size max, Rate (as decimal or percentage)
    """
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        lo = _safe_float(row[0])
        hi = _safe_float(row[1])
        rate = _safe_float(row[2])
        if lo is None or hi is None or rate is None:
            continue
        # Normalise percentage stored as whole number
        if rate > 1:
            rate = rate / 100
        rows.append({"size_min_mwp": lo, "size_max_mwp": hi, "rate_pct": rate})
    return rows


def parse_state_tax(ws) -> list[dict]:
    """
    Sales & Use Tax sheet.
    Columns: State code, State name, Base rate %, Solar exempt (Y/N), Notes
    """
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        code = str(row[0]).strip().upper() if row[0] else None
        if not code or len(code) != 2:
            continue
        rate = _safe_float(row[2], 0)
        if rate > 1:
            rate = rate / 100
        exempt_raw = str(row[3]).strip().upper() if row[3] else "N"
        rows.append({
            "state_code":    code,
            "state_name":    str(row[1]).strip() if row[1] else code,
            "base_rate_pct": rate,
            "solar_exempt":  exempt_raw in ("Y", "YES", "TRUE", "1"),
            "notes":         str(row[4]).strip() if row[4] else None,
        })
    return rows


# ─────────────────────────────────────────────────────────────────────────────
# DB WRITE HELPERS
# ─────────────────────────────────────────────────────────────────────────────

def upsert_system_rates(cur, rows: list[dict]) -> int:
    cur.execute("TRUNCATE pricing_system_rates RESTART IDENTITY")
    for r in rows:
        cur.execute("""
            INSERT INTO pricing_system_rates
                (system_type, size_min_mwp, size_max_mwp,
                 module_per_wp, inverter_per_wp, racking_per_wp, racking_sat_per_wp,
                 bos_per_wp, mechanical_per_wp, electrical_per_wp, civil_per_wp,
                 overhead_per_wp, sga_per_wp, contingency_pct, margin_pct, source)
            VALUES (%(system_type)s, %(size_min_mwp)s, %(size_max_mwp)s,
                    %(module_per_wp)s, %(inverter_per_wp)s, %(racking_per_wp)s,
                    %(racking_sat_per_wp)s, %(bos_per_wp)s,
                    %(mechanical_per_wp)s, %(electrical_per_wp)s, %(civil_per_wp)s,
                    %(overhead_per_wp)s, %(sga_per_wp)s,
                    %(contingency_pct)s, %(margin_pct)s, 'excel_import')
        """, r)
    return len(rows)


def upsert_engineering(cur, rows: list[dict]) -> int:
    cur.execute("TRUNCATE pricing_engineering_fixed RESTART IDENTITY")
    for r in rows:
        cur.execute("""
            INSERT INTO pricing_engineering_fixed
                (size_min_mwp, size_max_mwp, electrical_usd, civil_usd, substation_usd, source)
            VALUES (%(size_min_mwp)s, %(size_max_mwp)s,
                    %(electrical_usd)s, %(civil_usd)s, %(substation_usd)s, 'excel_import')
        """, r)
    return len(rows)


def upsert_permitting(cur, rows: list[dict]) -> int:
    cur.execute("TRUNCATE pricing_permitting_fixed RESTART IDENTITY")
    for r in rows:
        cur.execute("""
            INSERT INTO pricing_permitting_fixed
                (size_min_mwp, size_max_mwp, total_usd,
                 local_counsel_usd, environmental_usd, civil_survey_usd, facilitation_usd, source)
            VALUES (%(size_min_mwp)s, %(size_max_mwp)s, %(total_usd)s,
                    %(local_counsel_usd)s, %(environmental_usd)s,
                    %(civil_survey_usd)s, %(facilitation_usd)s, 'excel_import')
        """, r)
    return len(rows)


def upsert_bonding(cur, rows: list[dict]) -> int:
    cur.execute("TRUNCATE pricing_bonding RESTART IDENTITY")
    for r in rows:
        cur.execute("""
            INSERT INTO pricing_bonding (size_min_mwp, size_max_mwp, rate_pct, source)
            VALUES (%(size_min_mwp)s, %(size_max_mwp)s, %(rate_pct)s, 'excel_import')
        """, r)
    return len(rows)


def upsert_state_tax(cur, rows: list[dict]) -> int:
    for r in rows:
        cur.execute("""
            INSERT INTO pricing_state_tax
                (state_code, state_name, base_rate_pct, solar_exempt, notes)
            VALUES (%(state_code)s, %(state_name)s, %(base_rate_pct)s,
                    %(solar_exempt)s, %(notes)s)
            ON CONFLICT (state_code) DO UPDATE SET
                state_name    = EXCLUDED.state_name,
                base_rate_pct = EXCLUDED.base_rate_pct,
                solar_exempt  = EXCLUDED.solar_exempt,
                notes         = EXCLUDED.notes,
                updated_at    = NOW()
        """, r)
    return len(rows)


# ─────────────────────────────────────────────────────────────────────────────
# MAIN
# ─────────────────────────────────────────────────────────────────────────────

def main(dry_run: bool = False):
    try:
        import openpyxl
    except ImportError:
        print("ERROR: openpyxl not installed. Run: uv add openpyxl")
        sys.exit(1)

    if not EXCEL_PATH.exists():
        print(f"ERROR: Excel file not found at:\n  {EXCEL_PATH}")
        sys.exit(1)

    print(f"Reading: {EXCEL_PATH.name}")
    wb = openpyxl.load_workbook(EXCEL_PATH, read_only=True, data_only=True)

    sheet_map = {s.title: s for s in wb.worksheets}
    print(f"Sheets found: {list(sheet_map.keys())}\n")

    # Parse — adjust sheet names if they differ in your Excel version
    system_rows  = parse_system_price(sheet_map.get("System Price", sheet_map.get("System Prices", next(iter(sheet_map.values())))))
    eng_rows     = parse_engineering(sheet_map["Engineering"]) if "Engineering" in sheet_map else []
    perm_rows    = parse_permitting(sheet_map["Permitting"]) if "Permitting" in sheet_map else []
    bond_rows    = parse_bonding(sheet_map["Bonding"]) if "Bonding" in sheet_map else []
    tax_rows     = parse_state_tax(sheet_map.get("Sales & Use Tax", sheet_map.get("Sales&Use Tax", {}.__class__()))) \
                   if any(k.startswith("Sales") for k in sheet_map) else []

    print("Parsed rows:")
    print(f"  System rates:  {len(system_rows)}")
    print(f"  Engineering:   {len(eng_rows)}")
    print(f"  Permitting:    {len(perm_rows)}")
    print(f"  Bonding:       {len(bond_rows)}")
    print(f"  State tax:     {len(tax_rows)}")

    if dry_run:
        print("\n--dry-run: no changes written to DB.")
        return

    conn = _get_conn()
    cur  = conn.cursor()
    try:
        n1 = upsert_system_rates(cur, system_rows) if system_rows else 0
        n2 = upsert_engineering(cur, eng_rows)      if eng_rows    else 0
        n3 = upsert_permitting(cur, perm_rows)      if perm_rows   else 0
        n4 = upsert_bonding(cur, bond_rows)         if bond_rows   else 0
        n5 = upsert_state_tax(cur, tax_rows)        if tax_rows    else 0

        conn.commit()
        print("\nDB updated:")
        print(f"  pricing_system_rates:     {n1} rows")
        print(f"  pricing_engineering_fixed:{n2} rows")
        print(f"  pricing_permitting_fixed: {n3} rows")
        print(f"  pricing_bonding:          {n4} rows")
        print(f"  pricing_state_tax:        {n5} rows (upsert)")
        print("\nDone. Benchmark log and session tables were not touched.")
    except Exception as e:
        conn.rollback()
        print(f"ERROR: {e}")
        raise
    finally:
        cur.close()
        conn.close()


if __name__ == "__main__":
    dry = "--dry-run" in sys.argv
    main(dry_run=dry)
