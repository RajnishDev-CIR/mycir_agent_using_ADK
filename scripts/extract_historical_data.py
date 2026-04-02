"""
Extract historical CAPEX data from CIR Excel files.

Usage:
    uv run python scripts/extract_historical_data.py

Reads:
  - cir_old_manual_flow_docs/capex/251217_Indicative Capex Check List_V0_SS_SV.xlsx
  - cir_old_manual_flow_docs/capex/IGS/*.xlsx

Outputs:
  - scripts/historical_data_extract.json   (machine-readable for calibration)
  - scripts/historical_data_extract.txt    (human-readable summary)
"""
from __future__ import annotations

import json
import re
import sys
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print("openpyxl not found. Run: uv add openpyxl")
    sys.exit(1)

ROOT = Path(__file__).parent.parent
CHECKLIST_PATH = ROOT / "cir_old_manual_flow_docs" / "capex" / "251217_Indicative Capex Check List_V0_SS_SV.xlsx"
IGS_DIR = ROOT / "cir_old_manual_flow_docs" / "capex" / "IGS"
OUT_JSON = Path(__file__).parent / "historical_data_extract.json"
OUT_TXT  = Path(__file__).parent / "historical_data_extract.txt"


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def cell_val(ws, row: int, col: int):
    """Return cell value, stripping whitespace from strings."""
    v = ws.cell(row=row, column=col).value
    if isinstance(v, str):
        v = v.strip()
    return v


def find_row_containing(ws, text: str, col: int = 1, max_row: int = 200) -> int | None:
    """Return first row index where column `col` cell contains `text` (case-insensitive)."""
    text_l = text.lower()
    for r in range(1, max_row + 1):
        v = ws.cell(row=r, column=col).value
        if v and text_l in str(v).lower():
            return r
    return None


def to_float(v) -> float | None:
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip().replace(",", "").replace("$", "").replace("%", "")
    if not s:
        return None
    try:
        return float(s)
    except ValueError:
        return None


def to_dollar_per_wp(usd: float | None, dc_wp: float) -> float | None:
    if usd is None or dc_wp == 0:
        return None
    return round(usd / dc_wp, 4)


# ---------------------------------------------------------------------------
# Checklist parser
# ---------------------------------------------------------------------------

def parse_checklist(path: Path) -> list[dict]:
    """
    Parse every sheet in the checklist workbook.
    Returns list of project input dicts.
    """
    wb = openpyxl.load_workbook(path, data_only=True)
    projects = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        # Dump all non-empty cells to help locate fields
        cells: dict[str, str] = {}
        for row in ws.iter_rows(min_row=1, max_row=100, max_col=20):
            for cell in row:
                if cell.value not in (None, ""):
                    cells[f"R{cell.row}C{cell.column}"] = str(cell.value).strip()

        if not cells:
            continue

        # Print raw cell dump for each sheet so we can see the layout
        project_entry = {
            "sheet": sheet_name,
            "raw_cells": cells,
        }
        projects.append(project_entry)

    wb.close()
    return projects


# ---------------------------------------------------------------------------
# IGS CAPEX output parser
# ---------------------------------------------------------------------------

# Keywords that map to our standard line items
LINE_ITEM_KEYWORDS = {
    "module":        ["module", "panel", "solar panel", "pv module"],
    "inverter":      ["inverter"],
    "racking":       ["racking", "structure", "mounting", "tracker", "sat", "rack"],
    "bos":           ["bos", "balance of system", "balance-of-system", "cable", "conduit", "combiner", "dc wiring"],
    "mechanical":    ["mechanical install", "mechanical", "mounting install", "panel install"],
    "electrical":    ["electrical install", "electrical", "ac wiring", "ac install"],
    "civil":         ["civil", "grading", "earthwork", "site prep", "geotechnical"],
    "engineering":   ["engineering", "design", "e&p", "eng &"],
    "permitting":    ["permitting", "permit", "ahj", "interconnect fee"],
    "overhead":      ["overhead", "sga", "s,g&a", "general & admin", "g&a", "indirect"],
    "contingency":   ["contingency"],
    "bonding":       ["bonding", "bond", "payment bond", "performance bond"],
    "transformer":   ["transformer", "step-up", "padmount"],
    "margin":        ["margin", "profit", "markup", "fee"],
    "mobilization":  ["mobilization", "mobilisation", "mob"],
}


def classify_line_item(label: str) -> str | None:
    label_l = label.lower().strip()
    for category, keywords in LINE_ITEM_KEYWORDS.items():
        for kw in keywords:
            if kw in label_l:
                return category
    return None


def parse_igs_file(path: Path) -> dict | None:
    """
    Parse a single IGS CAPEX output Excel file.
    Returns structured project dict or None if parsing fails.
    """
    try:
        wb = openpyxl.load_workbook(path, data_only=True)
    except Exception as e:
        return {"error": str(e), "file": path.name}

    result = {
        "file": path.name,
        "is_ext": "_EXT_" in path.name.upper(),
        "sheets": [],
        "project_name": None,
        "location": None,
        "installation_type": None,
        "dc_kwp": None,
        "ac_kw": None,
        "line_items": {},
        "total_usd": None,
        "total_per_wp": None,
        "margin_pct": None,
        "contingency_pct": None,
        "raw_rows": [],
    }

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        result["sheets"].append(sheet_name)

        # Collect all non-empty rows for analysis
        sheet_rows = []
        for row in ws.iter_rows(min_row=1, max_row=200, max_col=15):
            row_data = []
            for cell in row:
                if cell.value is not None:
                    row_data.append({
                        "col": cell.column,
                        "row": cell.row,
                        "val": str(cell.value).strip() if isinstance(cell.value, str) else cell.value,
                    })
            if row_data:
                sheet_rows.append(row_data)
        result["raw_rows"].extend(sheet_rows[:80])  # first 80 non-empty rows per sheet

        # Try to extract key fields from this sheet
        for row in ws.iter_rows(min_row=1, max_row=200, max_col=15):
            cells_in_row = [(c.column, c.value) for c in row if c.value not in (None, "")]
            if not cells_in_row:
                continue

            row_num = row[0].row
            labels = [str(v).strip() for _, v in cells_in_row]
            full_row_text = " | ".join(labels)

            # Project name detection
            for col, val in cells_in_row:
                if val and "project" in str(val).lower() and result["project_name"] is None:
                    # next cell in same row might be the value
                    for col2, val2 in cells_in_row:
                        if col2 > col and val2 and "project" not in str(val2).lower():
                            result["project_name"] = str(val2).strip()
                            break

            # DC size
            for col, val in cells_in_row:
                if val and result["dc_kwp"] is None:
                    s = str(val).lower()
                    m = re.search(r"(\d+\.?\d*)\s*(kwp|kw|mwp|mw)", s)
                    if m and ("dc" in full_row_text.lower() or "size" in full_row_text.lower() or "capacity" in full_row_text.lower()):
                        num = float(m.group(1))
                        unit = m.group(2)
                        if "mwp" in unit or ("mw" in unit and "kwp" not in unit and "kw" not in unit):
                            result["dc_kwp"] = num * 1000
                        else:
                            result["dc_kwp"] = num

            # Try line-item detection — row has a label col + numeric value col(s)
            if len(cells_in_row) >= 2:
                first_label = str(cells_in_row[0][1]).strip()
                category = classify_line_item(first_label)
                if category:
                    # Find numeric values in the row
                    amounts = []
                    for col, val in cells_in_row[1:]:
                        f = to_float(val)
                        if f is not None and abs(f) > 100:  # skip small %/rate values
                            amounts.append((col, f))
                    if amounts:
                        # Take the largest amount as the USD value
                        _, usd = max(amounts, key=lambda x: abs(x[1]))
                        result["line_items"][category] = {
                            "usd": usd,
                            "label": first_label,
                            "row": row_num,
                        }

            # Total detection
            for i, (col, val) in enumerate(cells_in_row):
                if val and "total" in str(val).lower() and "epc" in str(val).lower():
                    for col2, val2 in cells_in_row[i+1:]:
                        f = to_float(val2)
                        if f and f > 10000:
                            result["total_usd"] = f
                            break

    wb.close()
    return result


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    all_output = {"checklist": [], "igs_projects": []}
    lines = []

    # --- Checklist ---
    lines.append("=" * 70)
    lines.append("CHECKLIST FILE: " + CHECKLIST_PATH.name)
    lines.append("=" * 70)
    if CHECKLIST_PATH.exists():
        cl_data = parse_checklist(CHECKLIST_PATH)
        all_output["checklist"] = cl_data
        for entry in cl_data:
            lines.append(f"\n--- Sheet: {entry['sheet']} ---")
            for k, v in entry["raw_cells"].items():
                lines.append(f"  {k}: {v}")
    else:
        lines.append("FILE NOT FOUND: " + str(CHECKLIST_PATH))

    # --- IGS files ---
    lines.append("\n\n" + "=" * 70)
    lines.append("IGS OUTPUT FILES")
    lines.append("=" * 70)

    if not IGS_DIR.exists():
        lines.append("IGS DIRECTORY NOT FOUND: " + str(IGS_DIR))
    else:
        igs_files = sorted(IGS_DIR.glob("*.xlsx"))
        lines.append(f"Found {len(igs_files)} files\n")

        for fp in igs_files:
            proj = parse_igs_file(fp)
            all_output["igs_projects"].append(proj)

            lines.append("\n" + "-" * 60)
            lines.append(f"FILE: {fp.name}")
            lines.append(f"Type: {'EXTERNAL (with margin)' if proj.get('is_ext') else 'INTERNAL (no margin)'}")
            lines.append(f"Sheets: {proj.get('sheets', [])}")
            lines.append(f"Project: {proj.get('project_name')}")
            lines.append(f"DC: {proj.get('dc_kwp')} kWp")
            lines.append(f"Total USD: {proj.get('total_usd')}")

            lines.append("Line items found:")
            for cat, data in proj.get("line_items", {}).items():
                lines.append(f"  {cat:20s}: ${data['usd']:>12,.0f}  (label: '{data['label']}')")

            lines.append("\nRAW ROWS (first 60 non-empty):")
            for row_cells in proj.get("raw_rows", [])[:60]:
                row_str = "  |  ".join(
                    f"C{c['col']}={repr(c['val'])}" for c in row_cells
                )
                lines.append(f"  R{row_cells[0]['row']}: {row_str}")

    # Write outputs
    OUT_JSON.write_text(json.dumps(all_output, indent=2, default=str), encoding="utf-8")
    OUT_TXT.write_text("\n".join(lines), encoding="utf-8")

    print(f"Done.")
    print(f"  Text summary : {OUT_TXT}")
    print(f"  JSON data    : {OUT_JSON}")
    print(f"\nIGS files parsed: {len(all_output['igs_projects'])}")


if __name__ == "__main__":
    main()
